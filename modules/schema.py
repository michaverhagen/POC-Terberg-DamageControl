""" This module generates a schema from a configuration file """

import os
import csv
import json
import fnmatch
import openpyxl

from modules.utilities import DEFAULT_TYPES


def _class_add_property(instance, newclass, sheet, row):

    prop = {}
    prop['description'] = str(sheet.cell(row=row, column=1).value)
    prop['name'] = str(sheet.cell(row=row, column=3).value)
    datatype = str(sheet.cell(row=row, column=4).value)
    prop['dataType'] = []
    prop['dataType'].append(datatype)
    prop['indexInverted'] = bool(sheet.cell(row=row, column=6).value)

    modulename = str(sheet.cell(row=row, column=5).value)
    if modulename == "none":
        pass
    else:
        if modulename == "config":
            module = "text2vec-contextionary"
            if 'module_name' in instance:
                module = instance['module_name']
        else:
            module = modulename
        prop['moduleConfig'] = {}
        prop['moduleConfig'][module] = {}
        prop['moduleConfig'][module]['skip'] = bool(sheet.cell(row=row, column=7).value)
        prop['moduleConfig'][module]['vectorizePropertyName'] = bool(sheet.cell(row=row, column=8).value)

    newclass['properties'].append(prop)


def _schema_add_class(instance, schema, sheet, row):

    classname = str(sheet.cell(row=row, column=2).value)
    modulename = str(sheet.cell(row=row, column=5).value)
    vectorizeClassName = bool(sheet.cell(row=row, column=9).value)

    found = False
    newclass = None
    for temp in schema['classes']:
        if temp['class'] == classname:
            newclass = temp
            found = True
            break

    if not found:
        newclass = {}
        newclass['class'] = classname

        if modulename == "none":
            newclass['vectorizer'] = "none"
            newclass['properties'] = []
        elif modulename == "img2vec-neural":
            newclass['vectorizer'] = modulename
            newclass['moduleConfig'] = {}
            newclass['moduleConfig'][modulename] = {}
            newclass['moduleConfig'][modulename]['imageFields'] = ["image"]
            newclass['properties'] = []
        else:
            if modulename == "config":
                module = "text2vec-contextionary"
                if 'module_name' in instance:
                    module = instance['module_name']
            else:
                module = modulename

            newclass['vectorizer'] = module
            newclass['moduleConfig'] = {}
            newclass['moduleConfig'][module] = {}
            newclass['moduleConfig'][module]['vectorizeClassName'] = vectorizeClassName
            newclass['properties'] = []

        schema['classes'].append(newclass)

    return newclass


def _read_schema_excel(instance: dict, sheet: openpyxl.worksheet, schema: dict) -> dict:

    done = False
    row = 2
    while not done:
        value = sheet.cell(row=row, column=1).value
        if value is None:
            done = True
        elif isinstance(value, str):
            newclass = _schema_add_class(instance, schema, sheet, row)
            if newclass is not None:
                _class_add_property(instance, newclass, sheet, row)
            else:
                print("should not happen")

        row += 1



def create_schema(instance, path):

    schema = {}
    schema['classes'] = []

    if path is not None and os.path.exists(path) and fnmatch.fnmatch(path, "*.xlsx"):
        workbook = openpyxl.load_workbook(path, data_only=True)
        if workbook is not None:
            sheet = workbook.active
            if sheet is not None:
                 _read_schema_excel(instance, sheet, schema)

    path = "./schema/schema.json"
    if 'schema' in instance:
        path = instance['schema']

    with open(path, 'w') as jsonfile:
        json.dump(schema, jsonfile, indent=4)

    return schema


def load_schema(instance: dict, client):

    path = "./schema/schema.json"
    if instance is not None and 'schema' in instance:
        path = instance['schema']

    if client.schema.contains():
        client.schema.delete_all()
    client.schema.create(path)
